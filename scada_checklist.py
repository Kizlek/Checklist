import json
import os
import sys
from copy import deepcopy
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk


# Add more built-in processes here.
# Each process has a Start checklist and an End checklist.
# Each checklist row can also have its own process note.
DEFAULT_PROCESSES = {
    "General Startup": {                        #Process 1
        "Start": [
            {
                "item": "Main power supply is ON",
                "process_note": "",
            },
            {"item": "PLC communication is healthy", "process_note": ""},
            {"item": "Emergency stop circuit is reset", "process_note": ""},
            {"item": "HMI/SCADA screen is responding", "process_note": ""},
            {"item": "No active critical alarms", "process_note": ""},
            {"item": "Pump status is normal", "process_note": ""},
            {"item": "Valve positions are correct", "process_note": ""},
            {"item": "Tank level is within operating range", "process_note": ""},
            {"item": "Pressure is within operating range", "process_note": ""},
            {"item": "Area is clean and safe for operation", "process_note": ""},
        ],
        "End": [
            {"item": "Process has been stopped from SCADA", "process_note": ""},
            {"item": "Equipment is in a safe final state", "process_note": ""},
            {"item": "No new critical alarms are active", "process_note": ""},
            {"item": "Final readings have been checked", "process_note": ""},
            {"item": "Shift handover notes are complete", "process_note": ""},
        ],
    },
    "Pump Start": {                        #Process 2
        "Start": [
            {"item": "Pump area is clear", "process_note": ""},
            {
                "item": "Suction valve is open",
                "process_note": "",
            },
            {"item": "Discharge valve is in the correct position", "process_note": ""},
            {"item": "Pump motor is available", "process_note": ""},
            {"item": "No active pump alarms", "process_note": ""},
            {"item": "Start command has been confirmed", "process_note": ""},
            {"item": "Flow is stable after start", "process_note": ""},
        ],
        "End": [
            {"item": "Pump stop command has been confirmed", "process_note": ""},
            {"item": "Pump motor status is stopped", "process_note": ""},
            {"item": "Flow has returned to zero or expected standby value", "process_note": ""},
            {"item": "Final pump alarms have been checked", "process_note": ""},
        ],
    },
    "Tank Transfer": {                        #Process 3
        "Start": [
            {"item": "Source tank level is sufficient", "process_note": ""},
            {
                "item": "Destination tank has available capacity",
                "process_note": "",
            },
            {"item": "Transfer route valves are aligned", "process_note": ""},
            {"item": "Pump or transfer device is available", "process_note": ""},
            {"item": "No active high-level or low-level alarms", "process_note": ""},
            {"item": "Transfer flow is stable", "process_note": ""},
        ],
        "End": [
            {"item": "Transfer has been stopped", "process_note": ""},
            {"item": "Final source tank level has been checked", "process_note": ""},
            {"item": "Final destination tank level has been checked", "process_note": ""},
            {"item": "Transfer route has been returned to normal", "process_note": ""},
            {"item": "Final levels have been recorded", "process_note": ""},
        ],
    },
}


APP_TITLE = "Process Checklist"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXPORT_FOLDER = os.path.join(SCRIPT_DIR, "checklist_exports")
CUSTOM_PROCESS_FILE = os.path.join(SCRIPT_DIR, "custom_processes.json")
SETUP_CONFIG_FILE = os.path.join(SCRIPT_DIR, "setup_config.json")
ADMIN_PASSWORD = "admin"
DEFAULT_MACHINE_PROCESS_MAP = {
    "MCVD": ["General Startup", "MCA", "Tank Transfer"],
    "Stolp": ["General Startup", "Pump Start"],
    "Plasma": ["General Startup", "Pump Start", "Tank Transfer"],
    "xxx": ["General Startup"],
}
# Add or change subprocess options here.
# These options are shown after the operator selects a machine and process.
DEFAULT_MACHINE_PROCESS_SUBPROCESS_MAP = {
    "MCVD": {
        "General Startup": ["Standard"],
        "MCA": ["Standard"],
        "Tank Transfer": ["Standard"],
    },
    "Stolp": {
        "General Startup": ["Cu", "Al", "Au", "Ag"],
        "Pump Start": ["Cu", "Al", "Au", "Ag"],
    },
    "Plasma": {
        "General Startup": ["Standard"],
        "Pump Start": ["Standard"],
        "Tank Transfer": ["Standard"],
    },
    "Test": {
        "Test": ["Test Test"],
    },
}


def _load_setup_config():
    machine_process_map = deepcopy(DEFAULT_MACHINE_PROCESS_MAP)
    subprocess_map = deepcopy(DEFAULT_MACHINE_PROCESS_SUBPROCESS_MAP)

    if not os.path.exists(SETUP_CONFIG_FILE):
        return machine_process_map, subprocess_map

    try:
        with open(SETUP_CONFIG_FILE, "r", encoding="utf-8") as config_file:
            saved = json.load(config_file)
    except (OSError, json.JSONDecodeError):
        return machine_process_map, subprocess_map

    saved_machines = saved.get("machine_process_map", {})
    if isinstance(saved_machines, dict):
        for machine, processes in saved_machines.items():
            if isinstance(processes, list):
                machine_process_map[str(machine)] = [
                    str(process).strip() for process in processes if str(process).strip()
                ]

    saved_subprocesses = saved.get("machine_process_subprocess_map", {})
    if isinstance(saved_subprocesses, dict):
        for machine, process_map in saved_subprocesses.items():
            if not isinstance(process_map, dict):
                continue
            subprocess_map[str(machine)] = {}
            for process, subprocesses in process_map.items():
                if isinstance(subprocesses, list):
                    subprocess_map[str(machine)][str(process)] = [
                        str(subprocess).strip()
                        for subprocess in subprocesses
                        if str(subprocess).strip()
                    ]

    return machine_process_map, subprocess_map


def _save_setup_config():
    try:
        with open(SETUP_CONFIG_FILE, "w", encoding="utf-8") as config_file:
            json.dump(
                {
                    "machine_process_map": MACHINE_PROCESS_MAP,
                    "machine_process_subprocess_map": MACHINE_PROCESS_SUBPROCESS_MAP,
                },
                config_file,
                indent=2,
            )
        return True
    except OSError as error:
        messagebox.showerror("Save failed", f"Could not save setup lists:\n{error}")
        return False


MACHINE_PROCESS_MAP, MACHINE_PROCESS_SUBPROCESS_MAP = _load_setup_config()


class ScadaChecklistApp(tk.Tk):
    def __init__(
        self,
        mode="admin",
        operator="",
        machine="",
        machine_subclass="",
        locked_process="",
        initial_theoretical_values=None,
        theoretical_values_locked=False,
        open_values_on_start=False,
    ):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1280x760")
        self.minsize(1100, 640)
        self.configure(bg="#111827")

        self.mode = mode
        self.machine_name = machine
        self.machine_subclass = machine_subclass
        self.locked_process = locked_process
        self.processes = self._load_processes()
        first_process = locked_process if locked_process in self.processes else next(iter(self.processes), "")

        self.operator_var = tk.StringVar(value=operator)
        self.process_var = tk.StringVar(value=first_process)
        self.stage_var = tk.StringVar(value="Start")
        self.header_title_var = tk.StringVar()
        self.popup_enabled_var = tk.BooleanVar(value=True)
        self.previous_process_name = first_process
        self.status_var = tk.StringVar(value="READY")
        self.time_var = tk.StringVar()
        self.opened_time_var = tk.StringVar()
        self.selected_item_var = tk.IntVar(value=-1)
        self.checklist_opened_at = datetime.now()
        self.check_vars = []
        self.note_widgets = []
        self.changed_vars = []
        self.row_widgets = []
        self.process_note_labels = []
        self.checklist_state = {}
        self.theoretical_values = initial_theoretical_values or self._blank_value_rows()
        self.theoretical_values_locked = theoretical_values_locked
        self.actual_values = self._blank_value_rows()
        self.process_comment_text = ""
        self.values_window = None
        self.values_theoretical_entries = []
        self.values_actual_entries = []
        self.values_comment_widget = None
        self.open_values_on_start = open_values_on_start
        self.settings_dirty = False

        self._configure_styles()
        self._build_layout()
        self.protocol("WM_DELETE_WINDOW", self.close_app)
        self._load_selected_process()
        self._refresh_clock()
        if self.open_values_on_start:
            self.after(300, self.open_process_values_window)

    def _load_processes(self):
        processes = self._normalize_processes(deepcopy(DEFAULT_PROCESSES))

        if not os.path.exists(CUSTOM_PROCESS_FILE):
            return processes

        try:
            with open(CUSTOM_PROCESS_FILE, "r", encoding="utf-8") as process_file:
                saved_processes = json.load(process_file)
        except (OSError, json.JSONDecodeError):
            return processes

        if isinstance(saved_processes, dict) and "processes" in saved_processes:
            saved_processes = saved_processes.get("processes", {})
            saved_processes = self._normalize_processes(saved_processes)
        else:
            saved_processes = self._normalize_processes(saved_processes)

        if isinstance(saved_processes, dict):
            processes.update(saved_processes)

        return processes

    def _normalize_processes(self, raw_processes):
        normalized = {}

        if not isinstance(raw_processes, dict):
            return normalized

        for name, data in raw_processes.items():
            process_name = str(name).strip()
            if not process_name:
                continue

            if isinstance(data, list):
                normalized[process_name] = {
                    "Start": self._normalize_entries(data),
                    "End": self._default_end_entries(),
                }
                continue

            if not isinstance(data, dict):
                continue

            if "Start" in data or "End" in data:
                normalized[process_name] = {
                    "Start": self._normalize_entries(data.get("Start", [])),
                    "End": self._normalize_entries(data.get("End", []), ""),
                }
                if not normalized[process_name]["Start"]:
                    normalized[process_name]["Start"] = self._normalize_entries(data.get("items", []))
                if not normalized[process_name]["End"]:
                    normalized[process_name]["End"] = self._default_end_entries()
                continue

            normalized[process_name] = {
                "Start": self._normalize_entries(data.get("items", [])),
                "End": self._default_end_entries(),
            }

        return normalized

    def _normalize_entries(self, entries, default_note=""):
        normalized = []
        if not isinstance(entries, list):
            return normalized

        for entry in entries:
            if isinstance(entry, dict):
                item_text = str(entry.get("item", "")).strip()
                process_note = str(entry.get("process_note", entry.get("process_notes", default_note))).strip()
            else:
                item_text = str(entry).strip()
                process_note = str(default_note).strip()

            if item_text:
                normalized.append({"item": item_text, "process_note": process_note})

        return normalized

    def _default_end_entries(self):
        return [
            {"item": "Process has been ended safely", "process_note": ""},
            {"item": "Final alarms and readings have been checked", "process_note": ""},
        ]

    def _save_processes(self):
        try:
            with open(CUSTOM_PROCESS_FILE, "w", encoding="utf-8") as process_file:
                json.dump(
                    {
                        "processes": self.processes,
                    },
                    process_file,
                    indent=2,
                )
            return True
        except OSError as error:
            messagebox.showerror("Save failed", f"Could not save process list:\n{error}")
            return False

    def _configure_styles(self):
        self.style = ttk.Style(self)
        self.style.theme_use("clam")

        self.style.configure("Main.TFrame", background="#111827")
        self.style.configure("Panel.TFrame", background="#1f2937", relief="flat")
        self.style.configure(
            "Header.TLabel",
            background="#111827",
            foreground="#e5e7eb",
            font=("Segoe UI", 22, "bold"),
        )
        self.style.configure(
            "SubHeader.TLabel",
            background="#111827",
            foreground="#93c5fd",
            font=("Segoe UI", 11, "bold"),
        )
        self.style.configure(
            "Panel.TLabel",
            background="#1f2937",
            foreground="#e5e7eb",
            font=("Segoe UI", 10),
        )
        self.style.configure(
            "TableHead.TLabel",
            background="#374151",
            foreground="#f9fafb",
            font=("Segoe UI", 10, "bold"),
            padding=8,
        )
        self.style.configure(
            "Cell.TLabel",
            background="#1f2937",
            foreground="#e5e7eb",
            font=("Segoe UI", 10),
            padding=8,
        )
        self.style.configure(
            "Status.TLabel",
            background="#064e3b",
            foreground="#d1fae5",
            font=("Segoe UI", 13, "bold"),
            padding=10,
        )
        self.style.configure(
            "Danger.TLabel",
            background="#7f1d1d",
            foreground="#fee2e2",
            font=("Segoe UI", 13, "bold"),
            padding=10,
        )
        self.style.configure("TButton", font=("Segoe UI", 10, "bold"), padding=8)
        self.style.configure("Export.TButton", background="#2563eb", foreground="#ffffff")

    def _build_layout(self):
        main = ttk.Frame(self, style="Main.TFrame", padding=18)
        main.pack(fill="both", expand=True)

        header = ttk.Frame(main, style="Main.TFrame")
        header.pack(fill="x")

        title_group = ttk.Frame(header, style="Main.TFrame")
        title_group.pack(side="left", fill="x", expand=True)

        ttk.Label(title_group, textvariable=self.header_title_var, style="Header.TLabel").pack(anchor="w")
        ttk.Label(
            title_group,
            text="Operator process verification",
            style="SubHeader.TLabel",
        ).pack(anchor="w", pady=(2, 0))

        clock_panel = ttk.Frame(header, style="Panel.TFrame", padding=12)
        clock_panel.pack(side="right")
        ttk.Label(clock_panel, text="SYSTEM TIME", style="Panel.TLabel").pack(anchor="e")
        ttk.Label(clock_panel, textvariable=self.time_var, style="Panel.TLabel").pack(anchor="e")

        top_panel = ttk.Frame(main, style="Panel.TFrame", padding=14)
        top_panel.pack(fill="x", pady=(18, 12))

        ttk.Label(top_panel, text="Operator", style="Panel.TLabel").pack(side="left")
        if self.mode == "admin":
            operator_entry = ttk.Entry(
                top_panel,
                textvariable=self.operator_var,
                font=("Segoe UI", 12),
                width=22,
            )
            operator_entry.pack(side="left", padx=(10, 18))
            operator_entry.focus_set()
        else:
            ttk.Label(top_panel, textvariable=self.operator_var, style="Panel.TLabel").pack(
                side="left",
                padx=(10, 18),
            )

        if self.mode == "user":
            ttk.Label(top_panel, text="Machine", style="Panel.TLabel").pack(side="left")
            machine_label = self.machine_name
            if self.machine_subclass:
                machine_label = f"{self.machine_name} / {self.machine_subclass}"
            ttk.Label(top_panel, text=machine_label, style="Panel.TLabel").pack(
                side="left",
                padx=(10, 18),
            )

        ttk.Label(top_panel, text="Process", style="Panel.TLabel").pack(side="left")
        if self.mode == "admin":
            self.process_combo = ttk.Combobox(
                top_panel,
                textvariable=self.process_var,
                values=list(self.processes.keys()),
                state="readonly",
                width=26,
                font=("Segoe UI", 11),
            )
            self.process_combo.pack(side="left", padx=(10, 18))
            self.process_combo.bind("<<ComboboxSelected>>", self._on_process_changed)
        else:
            self.process_combo = None
            ttk.Label(top_panel, textvariable=self.process_var, style="Panel.TLabel").pack(
                side="left",
                padx=(10, 18),
            )

        ttk.Button(top_panel, text="Switch to End Checklist", command=self.toggle_stage).pack(
            side="left",
            padx=(0, 18),
        )
        self.stage_button = top_panel.winfo_children()[-1]

        ttk.Label(top_panel, text="Checklist Opened", style="Panel.TLabel").pack(side="left")
        ttk.Label(top_panel, textvariable=self.opened_time_var, style="Panel.TLabel").pack(
            side="left",
            padx=(10, 18),
        )

        self.status_label = ttk.Label(top_panel, textvariable=self.status_var, style="Status.TLabel")
        self.status_label.pack(side="right")

        editor_bar = ttk.Frame(main, style="Panel.TFrame", padding=10)
        editor_bar.pack(fill="x", pady=(0, 12))

        if self.mode == "admin":
            ttk.Button(editor_bar, text="Copy Process", command=self.copy_process).pack(side="left")
            ttk.Button(editor_bar, text="Add Item", command=self.add_item).pack(
                side="left",
                padx=(10, 0),
            )
            ttk.Button(editor_bar, text="Edit Selected Item", command=self.edit_selected_item).pack(
                side="left",
                padx=(10, 0),
            )
            ttk.Button(editor_bar, text="Delete Selected Item", command=self.delete_selected_item).pack(
                side="left",
                padx=(10, 0),
            )
            ttk.Button(editor_bar, text="Edit Item Process Note", command=self.edit_process_note).pack(
                side="left",
                padx=(10, 0),
            )
            ttk.Button(editor_bar, text="Save Process", command=self.save_process_settings).pack(
                side="left",
                padx=(10, 0),
            )
        ttk.Button(editor_bar, text="Process Values", command=self.open_process_values_window).pack(
            side="left",
            padx=(0 if self.mode == "user" else 10, 0),
        )
        ttk.Checkbutton(
            editor_bar,
            text="Show process note popups",
            variable=self.popup_enabled_var,
        ).pack(side="left", padx=(12, 0))

        ttk.Label(
            editor_bar,
            text=(
                "Click a row to select it. Double-click the checklist item text to edit it."
                if self.mode == "admin"
                else "Click a row to select it. Add operator notes directly in the table."
            ),
            style="Panel.TLabel",
        ).pack(side="right")

        table_panel = ttk.Frame(main, style="Panel.TFrame", padding=0)
        table_panel.pack(fill="both", expand=True)

        self.canvas = tk.Canvas(
            table_panel,
            bg="#1f2937",
            highlightthickness=0,
            borderwidth=0,
        )
        scrollbar = ttk.Scrollbar(table_panel, orient="vertical", command=self.canvas.yview)
        self.rows_frame = ttk.Frame(self.canvas, style="Panel.TFrame", padding=(0, 0, 0, 8))

        self.rows_frame.bind(
            "<Configure>",
            lambda event: self.canvas.configure(scrollregion=self.canvas.bbox("all")),
        )
        canvas_window = self.canvas.create_window((0, 0), window=self.rows_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar.set)
        self.canvas.bind(
            "<Configure>",
            lambda event: self.canvas.itemconfigure(canvas_window, width=event.width),
        )

        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        self._bind_checklist_scroll(self.canvas)
        self._bind_checklist_scroll(self.rows_frame)

        button_bar = ttk.Frame(main, style="Main.TFrame")
        button_bar.pack(fill="x", pady=(14, 0))

        ttk.Button(button_bar, text="Clear Checklist", command=self.clear_checklist).pack(side="left")
        ttk.Button(button_bar, text="Validate", command=self.validate_checklist).pack(
            side="left",
            padx=(10, 0),
        )
        close_wrapper = ttk.Frame(button_bar, style="Main.TFrame")
        close_wrapper.pack(side="left", fill="x", expand=True)
        ttk.Button(close_wrapper, text="Close", command=self.close_app).pack(anchor="center")
        ttk.Button(
            button_bar,
            text="Export to Excel",
            style="Export.TButton",
            command=self.export_to_excel,
        ).pack(side="right")

    def _active_process_name(self):
        return self.process_var.get().strip()

    def _active_stage(self):
        return self.stage_var.get()

    def _active_entries(self):
        process = self.processes.get(self._active_process_name(), {})
        if not isinstance(process, dict):
            return []
        return process.setdefault(self._active_stage(), [])

    def _active_items(self):
        return [entry["item"] for entry in self._active_entries()]

    def _selected_entry(self):
        item_index = self.selected_item_var.get()
        entries = self._active_entries()
        if 0 <= item_index < len(entries):
            return entries[item_index]
        return None

    def _entry_process_note(self, item_index):
        entries = self._active_entries()
        if 0 <= item_index < len(entries):
            return entries[item_index].get("process_note", "").strip()
        return ""

    def _mark_settings_dirty(self):
        self.settings_dirty = True
        self.status_var.set("UNSAVED SETTINGS")
        self.status_label.configure(style="Danger.TLabel")

    def _update_header_title(self):
        process = self._active_process_name() or "No process selected"
        stage = self._active_stage()
        if self.machine_name:
            machine_label = self.machine_name
            if self.machine_subclass:
                machine_label = f"{self.machine_name} / {self.machine_subclass}"
            title = f"{machine_label} - {process} - {stage} Checklist"
        else:
            title = f"{APP_TITLE} - {process} - {stage} Checklist"
        self.header_title_var.set(title)
        self.title(title)

    def save_process_settings(self):
        if self._save_processes():
            self.settings_dirty = False
            self.validate_checklist()
            messagebox.showinfo("Saved", "Process checklist settings saved.")

    def close_app(self):
        self._sync_process_values_from_window()
        if self.settings_dirty:
            answer = messagebox.askyesnocancel(
                "Unsaved process settings",
                "Process checklist settings were changed. Save them before closing?",
            )
            if answer is None:
                return
            if answer:
                if not self._save_processes():
                    return
        self.destroy()

    def toggle_stage(self):
        self._save_current_checklist_state()
        self.stage_var.set("End" if self._active_stage() == "Start" else "Start")
        self.stage_button.configure(
            text="Switch to Start Checklist"
            if self._active_stage() == "End"
            else "Switch to End Checklist"
        )
        self._update_header_title()
        self._load_selected_process(reset_opened_time=False)

    def _confirm_losing_work(self, title):
        if not self._has_started_checklist():
            return True

        return messagebox.askyesno(
            title,
            "Changing checklist clears current checks and operator notes. Continue?",
        )

    def _on_process_changed(self, _event=None):
        selected_process = self.process_var.get().strip()

        if not self._confirm_losing_work("Change process"):
            self.process_var.set(self.previous_process_name)
            self.process_combo.set(self.previous_process_name)
            return

        self.previous_process_name = selected_process
        self._update_header_title()
        self._load_selected_process()

    def _state_key(self):
        return (self._active_process_name(), self._active_stage())

    def _save_current_checklist_state(self):
        if not self.check_vars and not self.note_widgets:
            return

        self.checklist_state[self._state_key()] = {
            "checked": [done_var.get() for done_var in self.check_vars],
            "notes": [note.get("1.0", "end").strip() for note in self.note_widgets],
            "changed": [changed_var.get() for changed_var in self.changed_vars],
        }

    def _restore_current_checklist_state(self):
        saved = self.checklist_state.get(self._state_key())
        if not saved:
            return

        for index, value in enumerate(saved.get("checked", [])):
            if index < len(self.check_vars):
                self.check_vars[index].set(bool(value))

        for index, value in enumerate(saved.get("notes", [])):
            if index < len(self.note_widgets):
                self.note_widgets[index].delete("1.0", "end")
                self.note_widgets[index].insert("1.0", value)

        for index, value in enumerate(saved.get("changed", [])):
            if index < len(self.changed_vars):
                self.changed_vars[index].set(value)

    def _has_started_checklist(self):
        if any(done_var.get() for done_var in self.check_vars):
            return True
        return any(note.get("1.0", "end").strip() for note in self.note_widgets)

    def _load_selected_process(self, reset_opened_time=True, show_process_notes=True):
        self._update_header_title()
        if reset_opened_time:
            self.checklist_opened_at = datetime.now()
            self.opened_time_var.set(self._format_timestamp(self.checklist_opened_at))
        self.selected_item_var.set(-1)
        self._build_checklist_rows()
        self._restore_current_checklist_state()
        self.validate_checklist()
        if show_process_notes and self.popup_enabled_var.get():
            self._show_stage_notes_warning()

    def _show_stage_notes_warning(self):
        notes = []
        for entry in self._active_entries():
            note = entry.get("process_note", "").strip()
            if note:
                notes.append(f"- {entry['item']}: {note}")
        if notes:
            messagebox.showwarning("Process notes", "\n".join(notes))

    def _build_checklist_rows(self):
        for child in self.rows_frame.winfo_children():
            child.destroy()

        self.check_vars = []
        self.note_widgets = []
        self.changed_vars = []
        self.row_widgets = []
        self.process_note_labels = []

        headers = ["Done", "Checklist Item", "Process Notes", "Operator Notes", "Last Changed"]
        widths = [8, 42, 30, 34, 22]

        for column, header in enumerate(headers):
            label = ttk.Label(self.rows_frame, text=header, style="TableHead.TLabel", anchor="w")
            label.grid(row=0, column=column, sticky="ew", padx=(0, 1), pady=(0, 1))
            weight = 1 if column in (1, 2, 3) else 0
            self.rows_frame.grid_columnconfigure(column, weight=weight, minsize=widths[column] * 8)

        for index, entry in enumerate(self._active_entries(), start=1):
            self._build_checklist_row(index, entry)

        self.canvas.yview_moveto(0)

    def _build_checklist_row(self, row_number, entry):
        item_index = row_number - 1
        item = entry["item"]

        done_var = tk.BooleanVar(value=False)
        self.check_vars.append(done_var)
        changed_var = tk.StringVar(value="")
        self.changed_vars.append(changed_var)

        check_cell = tk.Frame(self.rows_frame, bg="#1f2937")
        check_cell.grid(row=row_number, column=0, sticky="nsew", padx=(0, 1), pady=(0, 1))
        check = tk.Checkbutton(
            check_cell,
            variable=done_var,
            bg="#1f2937",
            activebackground="#1f2937",
            selectcolor="#111827",
            fg="#e5e7eb",
            activeforeground="#e5e7eb",
            command=lambda: self._mark_item_changed(item_index, validate=True),
        )
        check.pack(padx=10, pady=8)

        item_label = tk.Label(
            self.rows_frame,
            text=item,
            anchor="w",
            justify="left",
            wraplength=500,
            bg="#1f2937",
            fg="#e5e7eb",
            font=("Segoe UI", 10),
            padx=8,
            pady=8,
        )
        item_label.grid(row=row_number, column=1, sticky="nsew", padx=(0, 1), pady=(0, 1))
        if self.mode == "admin":
            item_label.bind("<Double-Button-1>", lambda _event: self.edit_item(item_index))

        process_note_label = tk.Label(
            self.rows_frame,
            text=entry.get("process_note", ""),
            anchor="nw",
            justify="left",
            wraplength=280,
            bg="#1f2937",
            fg="#cbd5e1",
            font=("Segoe UI", 9),
            padx=8,
            pady=8,
        )
        process_note_label.grid(row=row_number, column=2, sticky="nsew", padx=(0, 1), pady=(0, 1))
        self.process_note_labels.append(process_note_label)

        note = tk.Text(
            self.rows_frame,
            height=2,
            wrap="word",
            font=("Segoe UI", 9),
            bg="#111827",
            fg="#f9fafb",
            insertbackground="#f9fafb",
            relief="flat",
            padx=8,
            pady=6,
        )
        note.grid(row=row_number, column=3, sticky="nsew", padx=(0, 1), pady=(0, 1))
        note.bind("<KeyRelease>", lambda _event: self._mark_item_changed(item_index))
        self.note_widgets.append(note)

        changed_label = tk.Label(
            self.rows_frame,
            textvariable=changed_var,
            anchor="w",
            justify="left",
            bg="#1f2937",
            fg="#cbd5e1",
            font=("Segoe UI", 9),
            padx=8,
            pady=8,
        )
        changed_label.grid(row=row_number, column=4, sticky="nsew", padx=(0, 1), pady=(0, 1))

        row_widgets = [check_cell, check, item_label, process_note_label, note, changed_label]
        self.row_widgets.append(row_widgets)
        for widget in row_widgets:
            widget.bind("<Button-1>", lambda _event, index=item_index: self.select_item(index), add="+")
            self._bind_checklist_scroll(widget)

    def _refresh_clock(self):
        self.time_var.set(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        self.after(1000, self._refresh_clock)

    def _format_timestamp(self, value):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    def select_item(self, item_index):
        was_selected = self.selected_item_var.get() == item_index
        self.selected_item_var.set(item_index)
        self._refresh_row_highlights()
        if not was_selected:
            self._maybe_show_item_process_note(item_index)

    def _refresh_row_highlights(self):
        for index, widgets in enumerate(self.row_widgets):
            selected = index == self.selected_item_var.get()
            background = "#334155" if selected else "#1f2937"
            note_background = "#243447" if selected else "#111827"

            for widget in widgets:
                if isinstance(widget, tk.Text):
                    widget.configure(bg=note_background)
                else:
                    widget.configure(bg=background)

    def _mark_item_changed(self, item_index, validate=False):
        if 0 <= item_index < len(self.changed_vars):
            self.changed_vars[item_index].set(self._format_timestamp(datetime.now()))
        if validate:
            self.validate_checklist()

    def _maybe_show_item_process_note(self, item_index):
        if not self.popup_enabled_var.get():
            return

        note = self._entry_process_note(item_index)
        if note:
            messagebox.showwarning("Process notes", note)

    def _bind_checklist_scroll(self, widget):
        widget.bind("<MouseWheel>", self._on_checklist_scroll, add="+")
        widget.bind("<Button-4>", self._on_checklist_scroll, add="+")
        widget.bind("<Button-5>", self._on_checklist_scroll, add="+")

    def _on_checklist_scroll(self, event):
        if getattr(event, "num", None) == 4:
            self.canvas.yview_scroll(-1, "units")
        elif getattr(event, "num", None) == 5:
            self.canvas.yview_scroll(1, "units")
        else:
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        return "break"

    def _get_results(self):
        timestamp = datetime.now()
        rows = []

        for item, done_var, note_widget in zip(
            self._active_entries(),
            self.check_vars,
            self.note_widgets,
        ):
            item_index = len(rows)
            rows.append(
                {
                    "Timestamp": timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                    "Checklist Opened": self.opened_time_var.get(),
                    "Operator": self.operator_var.get().strip(),
                    "Machine": self.machine_name,
                    "Subprocess Class": self.machine_subclass,
                    "Process": self._active_process_name(),
                    "Checklist Stage": self._active_stage(),
                    "Process Notes": item.get("process_note", "").strip(),
                    "Checklist Item": item["item"],
                    "Completed": "Yes" if done_var.get() else "No",
                    "Item Last Changed": self.changed_vars[item_index].get(),
                    "Notes": note_widget.get("1.0", "end").strip(),
                }
            )

        return rows

    def _blank_value_rows(self):
        return [{"parameter": "", "value": ""} for _ in range(8)]

    def open_process_values_window(self):
        if self.values_window and self.values_window.winfo_exists():
            self.values_window.lift()
            self.values_window.focus_set()
            return

        dialog = tk.Toplevel(self)
        self.values_window = dialog
        dialog.title("Process Values")
        dialog.transient(self)
        dialog.configure(bg="#111827")
        dialog.geometry("760x720")
        dialog.minsize(660, 600)

        main = ttk.Frame(dialog, style="Main.TFrame", padding=16)
        main.pack(fill="both", expand=True)
        main.grid_rowconfigure(0, weight=1)
        main.grid_columnconfigure(0, weight=1)

        value_canvas = tk.Canvas(main, bg="#111827", highlightthickness=0, borderwidth=0)
        value_scrollbar = ttk.Scrollbar(main, orient="vertical", command=value_canvas.yview)
        content = ttk.Frame(value_canvas, style="Main.TFrame")
        content_window = value_canvas.create_window((0, 0), window=content, anchor="nw")
        value_canvas.configure(yscrollcommand=value_scrollbar.set)

        content.bind(
            "<Configure>",
            lambda event: value_canvas.configure(scrollregion=value_canvas.bbox("all")),
        )
        value_canvas.bind(
            "<Configure>",
            lambda event: value_canvas.itemconfigure(content_window, width=event.width),
        )
        value_canvas.grid(row=0, column=0, sticky="nsew")
        value_scrollbar.grid(row=0, column=1, sticky="ns")
        self._bind_process_values_scroll(value_canvas, value_canvas)
        self._bind_process_values_scroll(content, value_canvas)

        ttk.Label(
            content,
            text="Theoretical / Wanted Set Values",
            style="SubHeader.TLabel",
        ).pack(anchor="w")
        theoretical_entries = self._build_value_table(
            content,
            self.theoretical_values,
            locked=self.theoretical_values_locked,
        )
        self.values_theoretical_entries = theoretical_entries

        ttk.Label(
            content,
            text="Actual Values",
            style="SubHeader.TLabel",
        ).pack(anchor="w", pady=(18, 0))
        actual_entries = self._build_value_table(content, self.actual_values)
        self.values_actual_entries = actual_entries

        ttk.Label(
            content,
            text="Subprocess Comments",
            style="SubHeader.TLabel",
        ).pack(anchor="w", pady=(18, 6))
        comments = tk.Text(
            content,
            height=7,
            wrap="word",
            font=("Segoe UI", 10),
            bg="#ffffff",
            fg="#111827",
            insertbackground="#111827",
            relief="solid",
            borderwidth=1,
            padx=8,
            pady=8,
        )
        comments.pack(fill="both", expand=True)
        comments.insert("1.0", self.process_comment_text)
        self.values_comment_widget = comments
        self._bind_process_values_scroll(comments, value_canvas)

        button_bar = ttk.Frame(main, style="Main.TFrame")
        button_bar.grid(row=1, column=0, sticky="ew", pady=(14, 0))

        def save_values():
            self._sync_process_values_from_window()

        def close_values():
            self._sync_process_values_from_window()
            self.values_window = None
            self.values_theoretical_entries = []
            self.values_actual_entries = []
            self.values_comment_widget = None
            dialog.destroy()

        ttk.Button(button_bar, text="Close", command=close_values).pack(side="right")
        ttk.Button(button_bar, text="Save Values", command=save_values).pack(
            side="right",
            padx=(0, 10),
        )

        self._bind_process_values_scroll_tree(main, value_canvas)
        self._bind_process_values_scroll_tree(content, value_canvas)
        dialog.protocol("WM_DELETE_WINDOW", close_values)

    def _bind_process_values_scroll(self, widget, canvas):
        widget.bind("<MouseWheel>", lambda event: self._on_process_values_scroll(event, canvas), add="+")
        widget.bind("<Button-4>", lambda event: self._on_process_values_scroll(event, canvas), add="+")
        widget.bind("<Button-5>", lambda event: self._on_process_values_scroll(event, canvas), add="+")

    def _bind_process_values_scroll_tree(self, widget, canvas):
        self._bind_process_values_scroll(widget, canvas)
        for child in widget.winfo_children():
            self._bind_process_values_scroll_tree(child, canvas)

    def _on_process_values_scroll(self, event, canvas):
        if not canvas.winfo_exists():
            return "break"

        if getattr(event, "num", None) == 4:
            canvas.yview_scroll(-1, "units")
        elif getattr(event, "num", None) == 5:
            canvas.yview_scroll(1, "units")
        else:
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        return "break"

    def _sync_process_values_from_window(self):
        if self.values_theoretical_entries:
            self.theoretical_values = self._read_value_entries(self.values_theoretical_entries)
        if self.values_actual_entries:
            self.actual_values = self._read_value_entries(self.values_actual_entries)
        if self.values_comment_widget and self.values_comment_widget.winfo_exists():
            self.process_comment_text = self.values_comment_widget.get("1.0", "end").strip()

    def _build_value_table(self, parent, values, locked=False):
        table = ttk.Frame(parent, style="Panel.TFrame", padding=8)
        table.pack(fill="x", pady=(8, 0))

        ttk.Label(table, text="Parameter", style="TableHead.TLabel").grid(
            row=0,
            column=0,
            sticky="ew",
            padx=(0, 1),
        )
        ttk.Label(table, text="Value", style="TableHead.TLabel").grid(
            row=0,
            column=1,
            sticky="ew",
        )
        table.grid_columnconfigure(0, weight=1)
        table.grid_columnconfigure(1, weight=1)

        entries = []
        for row_number, row in enumerate(values, start=1):
            parameter = ttk.Entry(table, font=("Segoe UI", 10))
            parameter.insert(0, row.get("parameter", ""))
            if locked:
                parameter.configure(state="disabled")
            parameter.grid(row=row_number, column=0, sticky="ew", padx=(0, 1), pady=(1, 0))

            value = ttk.Entry(table, font=("Segoe UI", 10))
            value.insert(0, row.get("value", ""))
            if locked:
                value.configure(state="disabled")
            value.grid(row=row_number, column=1, sticky="ew", pady=(1, 0))
            entries.append((parameter, value))

        return entries

    def _read_value_entries(self, entries):
        rows = []
        for parameter_entry, value_entry in entries:
            rows.append(
                {
                    "parameter": parameter_entry.get().strip(),
                    "value": value_entry.get().strip(),
                }
            )
        return rows

    def _refresh_process_combo(self):
        if self.process_combo is None:
            return
        self.process_combo.configure(values=list(self.processes.keys()))
        self.process_combo.set(self._active_process_name())

    def _ask_item_text(self, title, prompt, initial_value=""):
        dialog = tk.Toplevel(self)
        dialog.title(title)
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(bg="#1f2937")
        dialog.geometry("760x420")
        dialog.minsize(620, 320)

        result = {"value": None}

        ttk.Label(dialog, text=prompt, style="Panel.TLabel").pack(
            anchor="w",
            padx=16,
            pady=(16, 8),
        )
        text_box = tk.Text(
            dialog,
            height=12,
            wrap="word",
            font=("Segoe UI", 11),
            bg="#111827",
            fg="#f9fafb",
            insertbackground="#f9fafb",
            relief="flat",
            padx=10,
            pady=10,
        )
        text_box.pack(fill="both", expand=True, padx=16, pady=(0, 12))
        text_box.insert("1.0", initial_value)
        text_box.focus_set()

        button_bar = ttk.Frame(dialog, style="Panel.TFrame")
        button_bar.pack(fill="x", padx=16, pady=(0, 16))

        def save():
            result["value"] = text_box.get("1.0", "end").strip()
            dialog.destroy()

        def cancel():
            dialog.destroy()

        ttk.Button(button_bar, text="Cancel", command=cancel).pack(side="right")
        ttk.Button(button_bar, text="Save", command=save).pack(side="right", padx=(0, 10))

        dialog.bind("<Control-Return>", lambda _event: save())
        self.wait_window(dialog)
        return result["value"]

    def copy_process(self):
        source_name = self._active_process_name()
        if not source_name:
            return

        new_name = simpledialog.askstring(
            "Copy process",
            "Enter the name for the copied process:",
            initialvalue=f"{source_name} Copy",
            parent=self,
        )
        if not new_name:
            return

        new_name = new_name.strip()
        if not new_name:
            return

        if new_name in self.processes:
            messagebox.showwarning("Name already exists", "Please choose a different process name.")
            return

        self.processes[new_name] = deepcopy(self.processes[source_name])
        self.process_var.set(new_name)
        self.previous_process_name = new_name
        self._mark_settings_dirty()
        self._refresh_process_combo()
        self._load_selected_process()

    def add_item(self):
        process_name = self._active_process_name()
        if not process_name:
            return

        item_text = self._ask_item_text("Add item", "Enter the new checklist item:")
        if not item_text or not item_text.strip():
            return

        self._active_entries().append({"item": item_text.strip(), "process_note": ""})
        self._mark_settings_dirty()
        self._load_selected_process(reset_opened_time=False, show_process_notes=False)

    def edit_selected_item(self):
        self.edit_item(self.selected_item_var.get())

    def edit_item(self, item_index):
        process_name = self._active_process_name()
        items = self._active_entries()

        if item_index < 0 or item_index >= len(items):
            messagebox.showinfo("Select item", "Select a checklist item first.")
            return

        new_text = self._ask_item_text(
            "Edit item",
            "Edit checklist item:",
            initial_value=items[item_index]["item"],
        )
        if not new_text or not new_text.strip():
            return

        items[item_index]["item"] = new_text.strip()
        self._mark_settings_dirty()
        self._load_selected_process(reset_opened_time=False, show_process_notes=False)
        self.selected_item_var.set(item_index)
        self._mark_item_changed(item_index)
        self._refresh_row_highlights()

    def edit_process_note(self):
        entry = self._selected_entry()
        if entry is None:
            messagebox.showinfo("Select item", "Select a checklist item first.")
            return

        new_notes = self._ask_item_text(
            "Edit item process note",
            "Process note shown before checking or commenting on this item:",
            initial_value=entry.get("process_note", ""),
        )
        if new_notes is None:
            return

        entry["process_note"] = new_notes.strip()
        for label_index, label in enumerate(self.process_note_labels):
            label.configure(text=self._entry_process_note(label_index))
        self._mark_settings_dirty()

    def delete_selected_item(self):
        process_name = self._active_process_name()
        items = self._active_entries()
        item_index = self.selected_item_var.get()

        if item_index < 0 or item_index >= len(items):
            messagebox.showinfo("Select item", "Select a checklist item first.")
            return

        if not messagebox.askyesno("Delete item", f"Delete this item?\n\n{items[item_index]['item']}"):
            return

        del items[item_index]
        self._mark_settings_dirty()
        self._load_selected_process(reset_opened_time=False, show_process_notes=False)

    def validate_checklist(self):
        total = len(self.check_vars)
        completed = sum(var.get() for var in self.check_vars)

        if total and completed == total:
            self.status_var.set(f"COMPLETE  {completed}/{total}")
            self.status_label.configure(style="Status.TLabel")
        else:
            self.status_var.set(f"PENDING  {completed}/{total}")
            self.status_label.configure(style="Danger.TLabel")

    def clear_checklist(self):
        if not messagebox.askyesno("Clear checklist", "Clear all checks and notes?"):
            return

        for done_var in self.check_vars:
            done_var.set(False)
        for index, note_widget in enumerate(self.note_widgets):
            note_widget.delete("1.0", "end")
            self._mark_item_changed(index)
        self.validate_checklist()

    def export_to_excel(self):
        operator = self.operator_var.get().strip()
        if not operator:
            messagebox.showwarning("Operator required", "Please enter the Operator name first.")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror(
                "Excel export is not installed",
                "Install the Excel package first:\n\npython -m pip install openpyxl",
            )
            return

        self._sync_process_values_from_window()
        self._save_current_checklist_state()
        rows = self._get_results()
        timestamp = datetime.now()
        checklist_closed = self._format_timestamp(timestamp)
        safe_operator = self._safe_filename_part(operator)
        safe_process = self._safe_filename_part(self._active_process_name())
        default_name = f"{timestamp.strftime('%Y%m%d')}-{safe_operator}-{safe_process}.xlsx"

        os.makedirs(EXPORT_FOLDER, exist_ok=True)
        default_path = os.path.abspath(os.path.join(EXPORT_FOLDER, default_name))

        file_path = filedialog.asksaveasfilename(
            title="Export checklist to Excel",
            initialfile=os.path.basename(default_path),
            initialdir=os.path.dirname(default_path),
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
        )

        if not file_path:
            return

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Checklist"

        sheet["A1"] = "SCADA Operator Checklist"
        sheet["A1"].font = Font(size=18, bold=True, color="FFFFFF")
        sheet["A1"].fill = PatternFill(fill_type="solid", fgColor="1F2937")
        sheet.merge_cells("A1:F1")

        summary_rows = [
            ("Checklist Opened", self.opened_time_var.get()),
            ("Checklist Closed", checklist_closed),
            ("Operator", operator),
            ("Machine", self.machine_name if self.machine_name else "Admin mode"),
            ("Subprocess Class", self.machine_subclass if self.machine_subclass else ""),
            ("Process", self._active_process_name()),
            ("Checklist Stage", self._active_stage()),
        ]

        for row_number, (label, value) in enumerate(summary_rows, start=3):
            sheet.cell(row=row_number, column=1, value=label)
            sheet.cell(row=row_number, column=2, value=value)

        for cell in sheet["A"][2:8]:
            cell.font = Font(bold=True)

        sheet["A3"].font = Font(size=14, bold=True)
        sheet["B3"].font = Font(size=14, bold=True)
        sheet["A4"].font = Font(size=14, bold=True)
        sheet["B4"].font = Font(size=14, bold=True)

        values_start_row = 10
        sheet.cell(row=values_start_row, column=1, value="Theoretical / Wanted Set Values")
        sheet.cell(row=values_start_row, column=1).font = Font(size=13, bold=True)

        value_header_row = values_start_row + 1
        sheet.cell(row=value_header_row, column=1, value="Parameter")
        sheet.cell(row=value_header_row, column=2, value="Value")
        for cell in sheet[value_header_row]:
            cell.fill = PatternFill(fill_type="solid", fgColor="374151")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        current_row = value_header_row + 1
        for value_row in self.theoretical_values:
            if value_row.get("parameter") or value_row.get("value"):
                sheet.cell(row=current_row, column=1, value=value_row.get("parameter", ""))
                sheet.cell(row=current_row, column=2, value=value_row.get("value", ""))
                current_row += 1

        if current_row == value_header_row + 1:
            sheet.cell(row=current_row, column=1, value="")
            sheet.cell(row=current_row, column=2, value="")
            current_row += 1

        current_row += 1
        sheet.cell(row=current_row, column=1, value="Actual Values")
        sheet.cell(row=current_row, column=1).font = Font(size=13, bold=True)
        current_row += 1
        actual_header_row = current_row
        sheet.cell(row=actual_header_row, column=1, value="Parameter")
        sheet.cell(row=actual_header_row, column=2, value="Value")
        for cell in sheet[actual_header_row]:
            cell.fill = PatternFill(fill_type="solid", fgColor="374151")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        current_row += 1
        for value_row in self.actual_values:
            if value_row.get("parameter") or value_row.get("value"):
                sheet.cell(row=current_row, column=1, value=value_row.get("parameter", ""))
                sheet.cell(row=current_row, column=2, value=value_row.get("value", ""))
                current_row += 1

        if current_row == actual_header_row + 1:
            sheet.cell(row=current_row, column=1, value="")
            sheet.cell(row=current_row, column=2, value="")
            current_row += 1

        current_row += 1
        sheet.cell(row=current_row, column=1, value="Subprocess Comments")
        sheet.cell(row=current_row, column=1).font = Font(size=13, bold=True)
        current_row += 1
        sheet.cell(row=current_row, column=1, value=self.process_comment_text)
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
        sheet.cell(row=current_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")

        table_start_row = current_row + 3
        headers = ["Process Notes", "Checklist Item", "Completed", "Item Last Changed", "Operator Notes"]
        for column, header in enumerate(headers, start=1):
            sheet.cell(row=table_start_row, column=column, value=header)

        for row in rows:
            sheet.append(
                [
                    row["Process Notes"],
                    row["Checklist Item"],
                    row["Completed"],
                    row["Item Last Changed"],
                    row["Notes"],
                ]
            )

        header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in sheet[table_start_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row in sheet.iter_rows(min_row=table_start_row + 1):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        column_widths = {
            "A": 48,
            "B": 58,
            "C": 16,
            "D": 24,
            "E": 52,
            "F": 20,
        }
        for column, width in column_widths.items():
            sheet.column_dimensions[column].width = width

        for column in range(1, 6):
            sheet.cell(row=table_start_row, column=column).alignment = Alignment(horizontal="center")
            sheet.column_dimensions[get_column_letter(column)].bestFit = True

        for row_number in range(3, 9):
            sheet.cell(row=row_number, column=1).alignment = Alignment(horizontal="left")
            sheet.cell(row=row_number, column=2).alignment = Alignment(horizontal="left")

        workbook.save(file_path)

        messagebox.showinfo("Export complete", f"Checklist exported:\n{file_path}")

    def _safe_filename_part(self, value):
        return "".join(
            character if character.isalnum() or character in ("-", "_", ".") else "_"
            for character in value
        )


class AccessModeWindow(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Select Access Mode")
        self.geometry("420x260")
        self.resizable(False, False)
        self.configure(bg="#111827")
        self.result = None

        self._configure_styles()
        self._build_layout()
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    def _configure_styles(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("Portal.TFrame", background="#111827")
        style.configure(
            "PortalTitle.TLabel",
            background="#111827",
            foreground="#e5e7eb",
            font=("Segoe UI", 18, "bold"),
        )
        style.configure(
            "Portal.TLabel",
            background="#111827",
            foreground="#cbd5e1",
            font=("Segoe UI", 10),
        )
        style.configure("TButton", font=("Segoe UI", 10, "bold"), padding=8)

    def _build_layout(self):
        main = ttk.Frame(self, style="Portal.TFrame", padding=22)
        main.pack(fill="both", expand=True)

        ttk.Label(main, text=APP_TITLE, style="PortalTitle.TLabel").pack(anchor="center")
        ttk.Label(main, text="Select how you want to open the checklist.", style="Portal.TLabel").pack(
            anchor="center",
            pady=(8, 24),
        )

        ttk.Button(main, text="Admin", command=self._admin_login).pack(fill="x", pady=(0, 10))
        ttk.Button(main, text="User", command=self._open_user_setup).pack(fill="x")

    def _admin_login(self):
        password = simpledialog.askstring("Admin password", "Enter admin password:", show="*", parent=self)
        if password == ADMIN_PASSWORD:
            self.result = {"mode": "admin"}
            self.destroy()
            return

        if password is not None:
            messagebox.showerror("Access denied", "Incorrect admin password.")

    def _open_user_setup(self):
        self.result = {"mode": "user"}
        self.destroy()


class UserSetupWindow(tk.Tk):
    def __init__(self, admin_mode=False):
        super().__init__()
        self.admin_mode = admin_mode
        self.title("Admin Checklist Setup" if admin_mode else "Checklist Setup")
        self.geometry("760x430" if admin_mode else "520x400")
        self.resizable(False, False)
        self.configure(bg="#111827")
        self.result = None

        self.operator_var = tk.StringVar()
        self.machine_var = tk.StringVar(value=next(iter(MACHINE_PROCESS_MAP), "MCVD"))
        self.subclass_var = tk.StringVar()
        self.process_var = tk.StringVar()
        self.available_processes = self._available_process_names()

        self._configure_styles()
        self._build_layout()
        self._refresh_process_options()
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    def _configure_styles(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("Setup.TFrame", background="#111827")
        style.configure(
            "SetupTitle.TLabel",
            background="#111827",
            foreground="#e5e7eb",
            font=("Segoe UI", 18, "bold"),
        )
        style.configure(
            "Setup.TLabel",
            background="#111827",
            foreground="#e5e7eb",
            font=("Segoe UI", 10),
        )
        style.configure("TButton", font=("Segoe UI", 10, "bold"), padding=8)

    def _build_layout(self):
        main = ttk.Frame(self, style="Setup.TFrame", padding=22)
        main.pack(fill="both", expand=True)

        ttk.Label(
            main,
            text="Admin Checklist Setup" if self.admin_mode else "User Checklist Setup",
            style="SetupTitle.TLabel",
        ).grid(
            row=0,
            column=0,
            columnspan=3,
            sticky="w",
            pady=(0, 20),
        )

        current_row = 1
        if not self.admin_mode:
            ttk.Label(main, text="Operator", style="Setup.TLabel").grid(row=current_row, column=0, sticky="w", pady=8)
            operator_entry = ttk.Entry(main, textvariable=self.operator_var, width=34, font=("Segoe UI", 11))
            operator_entry.grid(row=current_row, column=1, sticky="ew", pady=8)
            operator_entry.focus_set()
            current_row += 1

        ttk.Label(main, text="Machine", style="Setup.TLabel").grid(row=current_row, column=0, sticky="w", pady=8)
        self.machine_combo = ttk.Combobox(
            main,
            textvariable=self.machine_var,
            values=list(MACHINE_PROCESS_MAP.keys()),
            state="readonly",
            width=31,
            font=("Segoe UI", 11),
        )
        self.machine_combo.grid(row=current_row, column=1, sticky="ew", pady=8)
        self.machine_combo.bind("<<ComboboxSelected>>", lambda _event: self._refresh_process_options())
        if self.admin_mode:
            machine_buttons = ttk.Frame(main, style="Setup.TFrame")
            machine_buttons.grid(row=current_row, column=2, sticky="w", padx=(10, 0))
            ttk.Button(machine_buttons, text="Add", command=self._add_machine).pack(side="left")
            ttk.Button(machine_buttons, text="Edit", command=self._edit_machine).pack(side="left", padx=(6, 0))
        current_row += 1

        ttk.Label(main, text="Process", style="Setup.TLabel").grid(row=current_row, column=0, sticky="w", pady=8)
        self.process_combo = ttk.Combobox(
            main,
            textvariable=self.process_var,
            values=[],
            state="readonly",
            width=31,
            font=("Segoe UI", 11),
        )
        self.process_combo.grid(row=current_row, column=1, sticky="ew", pady=8)
        self.process_combo.bind("<<ComboboxSelected>>", lambda _event: self._refresh_subprocess_options())
        if self.admin_mode:
            process_buttons = ttk.Frame(main, style="Setup.TFrame")
            process_buttons.grid(row=current_row, column=2, sticky="w", padx=(10, 0))
            ttk.Button(process_buttons, text="Add", command=self._add_process).pack(side="left")
            ttk.Button(process_buttons, text="Edit", command=self._edit_process).pack(side="left", padx=(6, 0))
        current_row += 1

        ttk.Label(main, text="Subprocess Class", style="Setup.TLabel").grid(row=current_row, column=0, sticky="w", pady=8)
        self.subclass_combo = ttk.Combobox(
            main,
            textvariable=self.subclass_var,
            values=[],
            state="readonly",
            width=31,
            font=("Segoe UI", 11),
        )
        self.subclass_combo.grid(row=current_row, column=1, sticky="ew", pady=8)
        if self.admin_mode:
            subprocess_buttons = ttk.Frame(main, style="Setup.TFrame")
            subprocess_buttons.grid(row=current_row, column=2, sticky="w", padx=(10, 0))
            ttk.Button(subprocess_buttons, text="Add", command=self._add_subprocess).pack(side="left")
            ttk.Button(subprocess_buttons, text="Edit", command=self._edit_subprocess).pack(side="left", padx=(6, 0))
        current_row += 1

        main.grid_columnconfigure(1, weight=1)
        ttk.Button(
            main,
            text="Open Admin Checklist" if self.admin_mode else "Open Checklist",
            command=self._open_checklist,
        ).grid(
            row=current_row,
            column=0,
            columnspan=3,
            sticky="ew",
            pady=(24, 0),
        )

    def _available_process_names(self):
        names = set(DEFAULT_PROCESSES.keys())
        if os.path.exists(CUSTOM_PROCESS_FILE):
            try:
                with open(CUSTOM_PROCESS_FILE, "r", encoding="utf-8") as process_file:
                    saved = json.load(process_file)
            except (OSError, json.JSONDecodeError):
                saved = {}

            if isinstance(saved, dict):
                saved_processes = saved.get("processes", saved)
                if isinstance(saved_processes, dict):
                    names.update(str(name) for name in saved_processes.keys())

        return sorted(name for name in names if name)

    def _refresh_process_options(self):
        machine = self.machine_var.get()
        preferred = MACHINE_PROCESS_MAP.get(machine, [])
        processes = [name for name in preferred if name in self.available_processes]
        if not processes:
            processes = self.available_processes

        self.process_combo.configure(values=processes)
        self.process_var.set(processes[0] if processes else "")
        self._refresh_subprocess_options()

    def _refresh_subprocess_options(self):
        machine = self.machine_var.get()
        process = self.process_var.get()
        subclasses_by_process = MACHINE_PROCESS_SUBPROCESS_MAP.get(machine, {})
        subclasses = subclasses_by_process.get(process, ["Standard"])
        self.subclass_combo.configure(values=subclasses)
        self.subclass_var.set(subclasses[0] if subclasses else "")

    def _add_machine(self):
        machine = simpledialog.askstring("Add machine", "Enter machine name:", parent=self)
        if not machine or not machine.strip():
            return

        machine = machine.strip()
        if machine in MACHINE_PROCESS_MAP:
            messagebox.showwarning("Already exists", "That machine already exists.")
            return

        MACHINE_PROCESS_MAP[machine] = []
        MACHINE_PROCESS_SUBPROCESS_MAP[machine] = {}
        _save_setup_config()
        self.machine_combo.configure(values=list(MACHINE_PROCESS_MAP.keys()))
        self.machine_var.set(machine)
        self._refresh_process_options()

    def _edit_machine(self):
        old_name = self.machine_var.get().strip()
        if not old_name:
            return

        new_name = simpledialog.askstring(
            "Edit machine",
            "Edit machine name:",
            initialvalue=old_name,
            parent=self,
        )
        if not new_name or not new_name.strip() or new_name.strip() == old_name:
            return

        new_name = new_name.strip()
        if new_name in MACHINE_PROCESS_MAP:
            messagebox.showwarning("Already exists", "That machine already exists.")
            return

        MACHINE_PROCESS_MAP[new_name] = MACHINE_PROCESS_MAP.pop(old_name, [])
        MACHINE_PROCESS_SUBPROCESS_MAP[new_name] = MACHINE_PROCESS_SUBPROCESS_MAP.pop(old_name, {})
        _save_setup_config()
        self.machine_combo.configure(values=list(MACHINE_PROCESS_MAP.keys()))
        self.machine_var.set(new_name)
        self._refresh_process_options()

    def _add_process(self):
        machine = self.machine_var.get().strip()
        if not machine:
            return

        process = simpledialog.askstring("Add process", "Enter process name:", parent=self)
        if not process or not process.strip():
            return

        process = process.strip()
        processes = MACHINE_PROCESS_MAP.setdefault(machine, [])
        if process in processes:
            messagebox.showwarning("Already exists", "That process already exists for this machine.")
            return

        processes.append(process)
        MACHINE_PROCESS_SUBPROCESS_MAP.setdefault(machine, {})[process] = ["Standard"]
        self._ensure_process_template(process)
        _save_setup_config()
        self.available_processes = self._available_process_names()
        self._refresh_process_options()
        self.process_var.set(process)
        self._refresh_subprocess_options()

    def _edit_process(self):
        machine = self.machine_var.get().strip()
        old_name = self.process_var.get().strip()
        if not machine or not old_name:
            return

        new_name = simpledialog.askstring(
            "Edit process",
            "Edit process name:",
            initialvalue=old_name,
            parent=self,
        )
        if not new_name or not new_name.strip() or new_name.strip() == old_name:
            return

        new_name = new_name.strip()
        processes = MACHINE_PROCESS_MAP.setdefault(machine, [])
        if new_name in processes:
            messagebox.showwarning("Already exists", "That process already exists for this machine.")
            return

        MACHINE_PROCESS_MAP[machine] = [new_name if process == old_name else process for process in processes]
        subprocesses = MACHINE_PROCESS_SUBPROCESS_MAP.setdefault(machine, {})
        subprocesses[new_name] = subprocesses.pop(old_name, ["Standard"])
        self._ensure_process_template(new_name)
        _save_setup_config()
        self.available_processes = self._available_process_names()
        self._refresh_process_options()
        self.process_var.set(new_name)
        self._refresh_subprocess_options()

    def _add_subprocess(self):
        machine = self.machine_var.get().strip()
        process = self.process_var.get().strip()
        if not machine or not process:
            return

        subprocess = simpledialog.askstring(
            "Add subprocess class",
            "Enter subprocess class:",
            parent=self,
        )
        if not subprocess or not subprocess.strip():
            return

        subprocess = subprocess.strip()
        subprocesses = MACHINE_PROCESS_SUBPROCESS_MAP.setdefault(machine, {}).setdefault(process, [])
        if subprocess in subprocesses:
            messagebox.showwarning("Already exists", "That subprocess class already exists.")
            return

        subprocesses.append(subprocess)
        _save_setup_config()
        self._refresh_subprocess_options()
        self.subclass_var.set(subprocess)

    def _edit_subprocess(self):
        machine = self.machine_var.get().strip()
        process = self.process_var.get().strip()
        old_name = self.subclass_var.get().strip()
        if not machine or not process or not old_name:
            return

        new_name = simpledialog.askstring(
            "Edit subprocess class",
            "Edit subprocess class:",
            initialvalue=old_name,
            parent=self,
        )
        if not new_name or not new_name.strip() or new_name.strip() == old_name:
            return

        new_name = new_name.strip()
        subprocesses = MACHINE_PROCESS_SUBPROCESS_MAP.setdefault(machine, {}).setdefault(process, [])
        if new_name in subprocesses:
            messagebox.showwarning("Already exists", "That subprocess class already exists.")
            return

        MACHINE_PROCESS_SUBPROCESS_MAP[machine][process] = [
            new_name if subprocess == old_name else subprocess for subprocess in subprocesses
        ]
        _save_setup_config()
        self._refresh_subprocess_options()
        self.subclass_var.set(new_name)

    def _ensure_process_template(self, process_name):
        DEFAULT_PROCESSES.setdefault(
            process_name,
            {
                "Start": [
                    {"item": "Add start checklist item", "process_note": ""},
                ],
                "End": [
                    {"item": "Add end checklist item", "process_note": ""},
                ],
            },
        )

        saved = {}
        if os.path.exists(CUSTOM_PROCESS_FILE):
            try:
                with open(CUSTOM_PROCESS_FILE, "r", encoding="utf-8") as process_file:
                    saved = json.load(process_file)
            except (OSError, json.JSONDecodeError):
                saved = {}

        processes = saved.get("processes", {}) if isinstance(saved, dict) else {}
        if process_name not in processes:
            processes[process_name] = deepcopy(DEFAULT_PROCESSES[process_name])
            try:
                with open(CUSTOM_PROCESS_FILE, "w", encoding="utf-8") as process_file:
                    json.dump({"processes": processes}, process_file, indent=2)
            except OSError as error:
                messagebox.showerror("Save failed", f"Could not save new process:\n{error}")

    def _open_checklist(self):
        operator = self.operator_var.get().strip()
        machine = self.machine_var.get().strip()
        subclass = self.subclass_var.get().strip()
        process = self.process_var.get().strip()

        if not operator and not self.admin_mode:
            messagebox.showwarning("Operator required", "Please enter the Operator name.")
            return
        if not machine or not process:
            messagebox.showwarning("Selection required", "Please select a machine and process.")
            return

        self.result = {
            "mode": "admin" if self.admin_mode else "user",
            "operator": operator,
            "machine": machine,
            "subclass": subclass,
            "process": process,
        }
        self.destroy()


class DesiredValuesSetupWindow(tk.Tk):
    def __init__(self, machine, subclass, process):
        super().__init__()
        self.title("Desired Process Values")
        self.geometry("620x520")
        self.minsize(560, 460)
        self.configure(bg="#111827")
        self.machine = machine
        self.subclass = subclass
        self.process = process
        self.result = None
        self.entries = []

        self._configure_styles()
        self._build_layout()
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    def _configure_styles(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("Setup.TFrame", background="#111827")
        style.configure(
            "SetupTitle.TLabel",
            background="#111827",
            foreground="#e5e7eb",
            font=("Segoe UI", 18, "bold"),
        )
        style.configure(
            "Setup.TLabel",
            background="#111827",
            foreground="#e5e7eb",
            font=("Segoe UI", 10),
        )
        style.configure(
            "TableHead.TLabel",
            background="#374151",
            foreground="#f9fafb",
            font=("Segoe UI", 10, "bold"),
            padding=8,
        )
        style.configure("TButton", font=("Segoe UI", 10, "bold"), padding=8)

    def _build_layout(self):
        main = ttk.Frame(self, style="Setup.TFrame", padding=22)
        main.pack(fill="both", expand=True)

        ttk.Label(main, text="Desired Process Values", style="SetupTitle.TLabel").pack(anchor="w")
        ttk.Label(
            main,
            text=f"{self.machine} / {self.subclass} - {self.process}",
            style="Setup.TLabel",
        ).pack(anchor="w", pady=(4, 18))

        table = ttk.Frame(main, style="Setup.TFrame")
        table.pack(fill="both", expand=True)
        table.grid_columnconfigure(0, weight=1)
        table.grid_columnconfigure(1, weight=1)

        ttk.Label(table, text="Parameter", style="TableHead.TLabel").grid(
            row=0,
            column=0,
            sticky="ew",
            padx=(0, 1),
        )
        ttk.Label(table, text="Wanted Value", style="TableHead.TLabel").grid(
            row=0,
            column=1,
            sticky="ew",
        )

        for row_number in range(1, 9):
            parameter = ttk.Entry(table, font=("Segoe UI", 10))
            parameter.grid(row=row_number, column=0, sticky="ew", padx=(0, 1), pady=(1, 0))
            value = ttk.Entry(table, font=("Segoe UI", 10))
            value.grid(row=row_number, column=1, sticky="ew", pady=(1, 0))
            self.entries.append((parameter, value))

        button_bar = ttk.Frame(main, style="Setup.TFrame")
        button_bar.pack(fill="x", pady=(18, 0))
        ttk.Button(button_bar, text="Skip", command=self._skip).pack(side="left")
        ttk.Button(button_bar, text="Continue to Checklist", command=self._continue).pack(side="right")

    def _read_entries(self):
        rows = []
        for parameter, value in self.entries:
            rows.append(
                {
                    "parameter": parameter.get().strip(),
                    "value": value.get().strip(),
                }
            )
        return rows

    def _skip(self):
        self.result = []
        self.destroy()

    def _continue(self):
        self.result = self._read_entries()
        self.destroy()


def main():
    portal = AccessModeWindow()
    portal.mainloop()

    if not portal.result:
        return

    setup = UserSetupWindow(admin_mode=portal.result["mode"] == "admin")
    setup.mainloop()
    if not setup.result:
        return

    if setup.result["mode"] == "admin":
        app = ScadaChecklistApp(
            mode="admin",
            operator=setup.result["operator"],
            machine=setup.result["machine"],
            machine_subclass=setup.result["subclass"],
            locked_process=setup.result["process"],
        )
        app.mainloop()
        return

    desired_values = DesiredValuesSetupWindow(
        machine=setup.result["machine"],
        subclass=setup.result["subclass"],
        process=setup.result["process"],
    )
    desired_values.mainloop()
    if desired_values.result is None:
        return
    theoretical_values_locked = any(
        row.get("parameter") or row.get("value") for row in desired_values.result
    )

    app = ScadaChecklistApp(
        mode="user",
        operator=setup.result["operator"],
        machine=setup.result["machine"],
        machine_subclass=setup.result["subclass"],
        locked_process=setup.result["process"],
        initial_theoretical_values=desired_values.result,
        theoretical_values_locked=theoretical_values_locked,
        open_values_on_start=True,
    )
    app.mainloop()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit(0)
